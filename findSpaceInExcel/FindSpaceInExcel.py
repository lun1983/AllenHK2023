# !/usr/bin/env python
# -*-coding:utf-8 -*-

"""
# File       : findSpaceInExcel.py
# Time       ：2024/3/20 10:59
# Author     ：CBIC-IT-TEAM
# version    ：python 3.11
# Conda Env  : base v23.9.0
# Description：
"""
import os

import openpyxl

CHECK_FILE = []
CHECK_RESULT = []
RESULT_FILE = './check_result.txt'


def readCheckfileName():
    for filename in os.listdir('./'):
        if (filename.endswith('.xls') or filename.endswith('.xlsx')) and '~$' not in filename:
            CHECK_FILE.append('./' + filename)

    return


def wirte_output(CHECK_RESULT):
    with open(RESULT_FILE, 'w', encoding='utf-8') as file:
        for resultx in CHECK_RESULT:
            file.write(resultx + '\n')

    return


def checkExcelfile():
    for xfile in CHECK_FILE:
        work_book = openpyxl.load_workbook(xfile)
        for sheet_name in work_book.sheetnames:
            sheet = work_book[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    # print("XXX:" + str(cell.value))
                    if cell.value is not None and str(cell.value).strip() == "":
                        # errmsg = "Excel:" + xfile + " Sheet:" + sheet_name + " Position:" + str(row)  + " Has Space !"
                        errmsg = "[EXCEL:]" + xfile + " [SHEET:]" + sheet_name + " [POSITION:]" + str(cell.coordinate) +  " [HAS SPACE!]"
                        CHECK_RESULT.append(errmsg)

    return

def main():
    print('****** findSpaceInExcel start *****')
    readCheckfileName()
    checkExcelfile()
    wirte_output(CHECK_RESULT)
    print('****** findSpaceInExcel end *****')


if __name__ == "__main__":
    main()
